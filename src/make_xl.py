from openpyxl import Workbook
from es_connect import es_connect, m_es_connect
from db_connect import db_connect, m_db_connect


def make_xl(info, start_day, qry):

    write_wb = Workbook()
    write_ws = write_wb.create_sheet("종류별_"+str(start_day))

    write_ws['A1'] = 'NO.'
    write_ws['B1'] = 'Nudge 종류'
    write_ws['C1'] = 'Slot 수'
    write_ws['D1'] = '넛지 노출 수'
    write_ws['E1'] = '넛지 클릭 수'
    write_ws['F1'] = '클릭 전환율'
    write_ws['G1'] = '노출 STB 수'
    write_ws['H1'] = '클릭 STB 수'
    write_ws['I1'] = '이용율'

    es_data = es_connect(info, start_day, qry)
    db_data = db_connect(info, qry)

    list_all_data = []

    for data in es_data:
        re_data = {"nudge_type": data['key']}
        list_show = data['show']['buckets']

        # 노출, 클릭, 노출 stb, 클릭 stb 의 count
        for show_data in list_show:
            re_data[show_data["key"]] = show_data['doc_count']
            re_data[show_data["key"] + "_stb"] = show_data['stb']['value']

        list_all_data.append(re_data)

    xl_number = 2

    for z in list_all_data:
        z['number'] = xl_number - 1
        for d in db_data:
            if list(dict(d).keys())[0] in z['nudge_type']:
                z['nudge_name'] = d[list(dict(d).keys())[0]]
        if 'click.voice_text.button' in z:
            click_p = z['click.voice_text.button'] / z['page_show'] * 100
            use_p = z['click.voice_text.button_stb'] / z['page_show_stb'] * 100
            z['click_p'] = click_p
            z['use_p'] = use_p
        else:
            z['click_p'] = 0
            z['use_p'] = 0
            z['click.voice_text.button_stb'] = 0
            z['click.voice_text.button'] = 0
        xl_number += 1
        print(z)
        write_ws.append([z['number'],
                         z['nudge_name'],
                         "",
                         z['page_show'],
                         z['click.voice_text.button'],
                         z['click_p'],
                         z['page_show_stb'],
                         z['click.voice_text.button_stb'],
                         z['use_p']])

    write_wb.save("../xlsx/자동넛지_종류별성과_통계데이터_"+str(start_day) + '.xlsx')


def m_make_xl(info, start_day, qry):

    write_wb = Workbook()
    write_ws = write_wb.create_sheet("위치별_"+str(start_day))

    write_ws['A1'] = 'NO.'
    write_ws['B1'] = '구분'
    write_ws['C1'] = 'Nudge 위치'
    write_ws['D1'] = 'Slot 수'
    write_ws['E1'] = '넛지 노출 수'
    write_ws['F1'] = '넛지 클릭 수'
    write_ws['G1'] = '클릭 전환율'
    write_ws['H1'] = '노출 STB 수'
    write_ws['I1'] = '클릭 STB 수'
    write_ws['J1'] = '이용율'

    es_data = m_es_connect(info, start_day, qry)
    db_data = m_db_connect(info, qry)

    list_all_data = []

    for data in es_data:
        re_data = {"nudge_type": data['key']}
        list_show = data['show']['buckets']

        # 노출, 클릭, 노출 stb, 클릭 stb 의 count
        for show_data in list_show:
            re_data[show_data["key"]] = show_data['doc_count']
            re_data[show_data["key"] + "_stb"] = show_data['stb']['value']

        list_all_data.append(re_data)

    xl_number = 2

    for z in list_all_data:
        z['number'] = xl_number - 1
        for d in db_data:
            if list(dict(d).keys())[0] in z['nudge_type']:
                z['menu_id'] = list(dict(d).values())[0]

        if 'click.voice_text.button' in z:
            click_p = z['click.voice_text.button'] / z['page_show'] * 100
            use_p = z['click.voice_text.button_stb'] / z['page_show_stb'] * 100
            z['click_p'] = click_p
            z['use_p'] = use_p
        else:
            z['click_p'] = 0
            z['use_p'] = 0
            z['click.voice_text.button_stb'] = 0
            z['click.voice_text.button'] = 0
        xl_number += 1
        write_ws.append([z['number'],
                         "",
                         z['menu_id'],
                         "",
                         z['page_show'],
                         z['click.voice_text.button'],
                         z['click_p'],
                         z['page_show_stb'],
                         z['click.voice_text.button_stb'],
                         z['use_p']])

    write_wb.save("../xlsx/자동넛지_위치별성과_통계데이터_"+str(start_day) + '.xlsx')
